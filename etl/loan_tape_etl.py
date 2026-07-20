"""Loads loan tape ETL.

Reads sources/loan_tape.xlsx (sheets: Loan Tape, expected_payments, actual_payments)
and produces etl/loads_loan_tape.json with a 6-segment hierarchy:
Total Portfolio / Loads Finance BU x All / Active / Closed.

Methodology
-----------
- One row per trade order. Financed exposure = totalWithAdjustments (order value net
  of credit/debit notes). Outstanding receivable = totalDueAmount as reported;
  overDueDays is the servicer's days-past-due field.
- Active = totalDueAmount > 0. Closed = fully collected (totalDueAmount = 0).
- Collections = Settled rows in actual_payments (regular + advance).
- DPD buckets follow the Manual de Credito 2026 provisioning grid; the expected-loss
  provision applies 0% (<=60d), 0.5% (61-90), 10% (91-120), 25% (121-150),
  65% (151-180), 100% (>180) to the outstanding receivable.
- NPL follows the Loads definition (receivables overdue > 180 days); PAR30/60/90
  are also reported on the active receivable base.
- Cohorts: monthly issue cohorts; collection rate = cumulative Settled collections
  divided by financed exposure of the cohort.
"""
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
SRC = HERE.parent / "sources" / "loan_tape.xlsx"
OUT = HERE / "loads_loan_tape.json"

PROVISION_GRID = [  # (min_dpd, max_dpd, rate) per Manual de Credito 2026 §6.2
    (1, 30, 0.0), (31, 60, 0.0), (61, 90, 0.005), (91, 120, 0.10),
    (121, 150, 0.25), (151, 180, 0.65), (181, 10**9, 1.0),
]
DPD_BUCKETS = [
    ("Current", 0, 0), ("1–30 DPD", 1, 30), ("31–60 DPD", 31, 60),
    ("61–90 DPD", 61, 90), ("91–150 DPD", 91, 150), ("151–180 DPD", 151, 180),
    ("> 180 DPD", 181, 10**9),
]
SCORE_TIERS = [  # Manual de Credito 2026 §3.2
    ("AAA", 723, 1000), ("AA", 606, 722), ("A", 546, 605), ("B", 475, 545),
    ("C", 252, 474), ("D", 152, 251), ("E", 0, 151),
]


def clean(s):
    if s is None:
        return None
    s = str(s)
    # Repair common UTF-8-as-latin-1 mojibake in counterparty names
    fixes = {"√â": "É", "√ç": "Í", "√ì": "Ó", "√Å": "Á", "√ñ": "Ö", "√ú": "Ü",
             "√©": "é", "√≠": "í", "√≥": "ó", "√°": "á", "√±": "ñ", "√∫": "ú"}
    for k, v in fixes.items():
        s = s.replace(k, v)
    return re.sub(r"\s+", " ", s).strip()


def d(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def month(dt):
    return dt.strftime("%Y-%m") if dt else None


def sheet_rows(wb, name):
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    return [dict(zip(hdr, r)) for r in rows[1:] if r and r[0]]


def provision_rate(dpd):
    if dpd <= 0:
        return 0.0
    for lo, hi, rate in PROVISION_GRID:
        if lo <= dpd <= hi:
            return rate
    return 1.0


def score_tier(score):
    if score is None:
        return "n/a"
    for name, lo, hi in SCORE_TIERS:
        if lo <= score <= hi:
            return name
    return "n/a"


def load():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    orders_raw = sheet_rows(wb, "Loan Tape")
    expected = sheet_rows(wb, "expected_payments")
    actual = sheet_rows(wb, "actual_payments")

    pays = defaultdict(list)
    for p in actual:
        if p.get("paymentStatus") != "Settled":
            continue  # validation: only settled cash counts as collections
        pays[p["orderCode"]].append({
            "date": d(p["paymentDate"]), "type": clean(p["paymentType"]),
            "amount": float(p["totalWithTax"] or 0),
        })

    exp = defaultdict(list)
    for e in expected:
        exp[e["orderCode"]].append({
            "code": clean(e["dealPaymentCode"]), "amount": float(e["dealPaymentAmount"] or 0),
            "due": d(e["dealPaymentDueDate"]),
        })

    orders = []
    for o in orders_raw:
        code = o["orderCode"]
        issue, due = d(o["issueDate"]), d(o["orderDueDate"])
        financed = float(o["totalWithAdjustments"] or 0)
        outstanding = float(o["totalDueAmount"] or 0)
        dpd = int(o["overDueDays"] or 0)
        opays = sorted(pays.get(code, []), key=lambda p: p["date"])
        collected = sum(p["amount"] for p in opays)
        orders.append({
            "code": code, "issue": issue, "due": due,
            "bu": clean(o["businessUnitCode"]),
            "importer": clean(o["importerName"]), "importer_id": o["importerId"],
            "importer_country": clean(o["importerCountry"]),
            "supplier": clean(o["supplierName"]), "supplier_country": clean(o["supplierCountry"]),
            "incoterm": clean(o["incoterm"]),
            "gross": float(o["totalOrderAmount"] or 0),
            "financed": financed,
            "credit_note": float(o["totalCreditNote"] or 0),
            "debit_note": float(o["totalDebitNote"] or 0),
            "product_cost": float(o["productCost"] or 0),
            "logistic_cost": float(o["logisticCost"] or 0),
            "outstanding": outstanding, "dpd": dpd if outstanding > 0 else 0,
            "score": int(o["creditCcore"]) if o["creditCcore"] is not None else None,
            "tier": score_tier(o["creditCcore"]),
            "collected": collected, "payments": opays,
            "expected": sorted(exp.get(code, []), key=lambda e: e["due"] or date.min),
            "active": outstanding > 0,
            "tenor_days": (due - issue).days if issue and due else None,
        })
    return orders


def seg_metrics(orders, label, parent):
    n = len(orders)
    if n == 0:
        return None
    amt = sum(o["financed"] for o in orders)
    active = [o for o in orders if o["active"]]
    ob = sum(o["outstanding"] for o in active)
    tenors = [o["tenor_days"] for o in orders if o["tenor_days"] is not None]
    wa = lambda vals_w: (sum(v * w for v, w in vals_w) / sum(w for _, w in vals_w)) if vals_w and sum(w for _, w in vals_w) else None
    # Trade margin measured on the gross order value; adjustments (credit/debit
    # notes for volume, quality, FX and logistics true-ups) are reported separately.
    margins = [((o["gross"] - o["product_cost"] - o["logistic_cost"]) / o["gross"], o["gross"])
               for o in orders if o["gross"] > 0]
    adj = sum(o["credit_note"] - o["debit_note"] for o in orders)
    scores = [(o["score"], o["financed"]) for o in orders if o["score"]]
    wa_tenor = wa([(t, o["financed"]) for t, o in zip(tenors, [x for x in orders if x["tenor_days"] is not None])])

    # credit quality on the active receivable base
    dist = []
    for name, lo, hi in DPD_BUCKETS:
        rows = [o for o in active if lo <= o["dpd"] <= hi]
        b_ob = sum(o["outstanding"] for o in rows)
        dist.append({"bucket": name, "outstanding": round(b_ob, 2),
                     "pct_opb": round(b_ob / ob, 6) if ob else 0.0, "count": len(rows)})

    def par(th):
        rows = [o for o in active if o["dpd"] >= th]
        a = sum(o["outstanding"] for o in rows)
        return {"amount": round(a, 2), "pct": round(a / ob, 6) if ob else 0.0, "count": len(rows)}

    cur = [o for o in active if o["dpd"] == 0]
    provision = sum(o["outstanding"] * provision_rate(o["dpd"]) for o in active)

    # collections
    monthly = defaultdict(float)
    by_type = defaultdict(float)
    total_coll = 0.0
    for o in orders:
        for p in o["payments"]:
            monthly[month(p["date"])] += p["amount"]
            by_type["advance" if p["type"] == "advance" else "regular"] += p["amount"]
            total_coll += p["amount"]

    # origination + cohorts
    orig_m = defaultdict(float)
    coh = defaultdict(lambda: {"disbursed": 0.0, "collected": 0.0, "count": 0})
    for o in orders:
        m = month(o["issue"])
        orig_m[m] += o["financed"]
        coh[m]["disbursed"] += o["financed"]
        coh[m]["collected"] += o["collected"]
        coh[m]["count"] += 1

    status = defaultdict(int)
    for o in orders:
        if not o["active"]:
            status["Collected in full"] += 1
        elif o["dpd"] == 0:
            status["Open — current"] += 1
        elif o["dpd"] <= 30:
            status["Open — 1–30 DPD"] += 1
        elif o["dpd"] <= 90:
            status["Open — 31–90 DPD"] += 1
        elif o["dpd"] <= 180:
            status["Open — 91–180 DPD"] += 1
        else:
            status["Open — > 180 DPD"] += 1

    top_ticket = sorted((o["financed"] for o in active), reverse=True)
    top5_share = sum(top_ticket[:5]) / ob if ob else None

    return {
        "label": label, "parent": parent,
        "origination": {
            "count": n, "amount": round(amt, 2),
            "first": min(o["issue"] for o in orders).isoformat(),
            "last": max(o["issue"] for o in orders).isoformat(),
            "avg_ticket": round(amt / n, 2),
            "tenor_days": {"min": min(tenors) if tenors else None,
                            "max": max(tenors) if tenors else None,
                            "wa": round(wa_tenor, 1) if wa_tenor else None},
        },
        "outstanding": {
            "active_ob": round(ob, 2), "active_count": len(active),
            "avg_balance": round(ob / len(active), 2) if active else None,
        },
        "weighted_averages": {
            "gross_trade_margin": round(wa(margins), 4) if margins else None,
            "wa_tenor_days": round(wa_tenor, 1) if wa_tenor else None,
            "wa_credit_score": round(wa(scores), 0) if scores else None,
            "net_adjustments": round(adj, 2),
        },
        "credit_quality": {
            "dpd_distribution": dist,
            "current": {"amount": round(sum(o["outstanding"] for o in cur), 2),
                         "pct": round(sum(o["outstanding"] for o in cur) / ob, 6) if ob else 0.0},
            "par30": par(31), "par60": par(61), "par90": par(91), "npl": par(181),
        },
        "provision": {
            "expected_loss": round(provision, 2),
            "net_receivable": round(ob - provision, 2),
            "coverage_pct": round(provision / ob, 6) if ob else 0.0,
        },
        "collections": {
            "total": round(total_coll, 2),
            "by_type": {k: round(v, 2) for k, v in sorted(by_type.items())},
            "monthly": [{"month": m, "value": round(v, 2)} for m, v in sorted(monthly.items()) if m],
            "collection_rate": round(total_coll / amt, 6) if amt else None,
        },
        "origination_monthly": [{"month": m, "value": round(v, 2)} for m, v in sorted(orig_m.items()) if m],
        "cohorts": [{"cohort": m, "disbursed": round(c["disbursed"], 2), "collected": round(c["collected"], 2),
                      "count": c["count"],
                      "collection_rate": round(c["collected"] / c["disbursed"], 4) if c["disbursed"] else None}
                     for m, c in sorted(coh.items()) if m],
        "status_distribution": dict(sorted(status.items(), key=lambda kv: -kv[1])),
        "top5_ticket_share": round(top5_share, 6) if top5_share is not None else None,
    }


def concentration(orders):
    active = [o for o in orders if o["active"]]
    ob = sum(o["outstanding"] for o in active) or 1.0
    fin = sum(o["financed"] for o in orders) or 1.0

    def top(key, rows, weight, base, k=8):
        agg = defaultdict(float)
        cnt = defaultdict(int)
        for o in rows:
            agg[o[key]] += o[weight]
            cnt[o[key]] += 1
        out = sorted(agg.items(), key=lambda kv: -kv[1])
        return [{"name": nm, "amount": round(v, 2), "pct": round(v / base, 6), "count": cnt[nm]}
                for nm, v in out[:k]]

    tiers = defaultdict(lambda: {"amount": 0.0, "count": 0})
    for o in orders:
        tiers[o["tier"]]["amount"] += o["financed"]
        tiers[o["tier"]]["count"] += 1
    tier_order = [t[0] for t in SCORE_TIERS] + ["n/a"]

    return {
        "importers_active_ob": top("importer", active, "outstanding", ob),
        "importer_countries_financed": top("importer_country", orders, "financed", fin, 10),
        "supplier_countries_financed": top("supplier_country", orders, "financed", fin, 10),
        "business_units_financed": top("bu", orders, "financed", fin, 10),
        "score_tiers": [{"tier": t, "amount": round(tiers[t]["amount"], 2),
                          "pct": round(tiers[t]["amount"] / fin, 6), "count": tiers[t]["count"]}
                         for t in tier_order if t in tiers],
    }


def outstanding_series(orders, as_of):
    """Month-end receivable balance reconstructed from order lifecycles.

    Per order: +financed at issue, -payment at each Settled payment, and a plug at
    the terminal month so the ending balance ties to the reported totalDueAmount
    (absorbs credit/debit notes booked at close)."""
    events = defaultdict(float)
    for o in orders:
        events[month(o["issue"])] += o["financed"]
        bal = o["financed"]
        last = o["issue"]
        for p in o["payments"]:
            events[month(p["date"])] -= p["amount"]
            bal -= p["amount"]
            last = max(last, p["date"])
        terminal = max(last, o["due"] or last)
        if o["active"]:
            terminal = max(terminal, as_of)
        plug = o["outstanding"] - bal
        if abs(plug) > 0.005:
            events[month(min(terminal, as_of))] += plug
    months = sorted(m for m in events if m)
    out, run = [], 0.0
    for m in months:
        run += events[m]
        out.append({"month": m, "value": round(max(run, 0.0), 2)})
    return out


def main():
    orders = load()
    as_of = max([o["issue"] for o in orders]
                + [p["date"] for o in orders for p in o["payments"]]
                + [o["due"] for o in orders if o["due"]])

    fin_orders = [o for o in orders if o["bu"] == "loads_finance"]
    segs = {
        "total": seg_metrics(orders, "Total Portfolio", None),
        "total_active": seg_metrics([o for o in orders if o["active"]], "Total Portfolio — Active", "total"),
        "total_closed": seg_metrics([o for o in orders if not o["active"]], "Total Portfolio — Closed", "total"),
        "fin": seg_metrics(fin_orders, "Loads Finance BU", None),
        "fin_active": seg_metrics([o for o in fin_orders if o["active"]], "Loads Finance BU — Active", "fin"),
        "fin_closed": seg_metrics([o for o in fin_orders if not o["active"]], "Loads Finance BU — Closed", "fin"),
    }
    segs = {k: v for k, v in segs.items() if v}

    payload = {
        "as_of": as_of.isoformat(),
        "segments": segs,
        "concentration": concentration(orders),
        "outstanding_series": outstanding_series(orders, as_of),
    }
    OUT.write_text(json.dumps(payload, indent=1), encoding="utf-8")
    t = segs["total"]
    print(f"loan tape ETL ok — {t['origination']['count']} orders, "
          f"${t['origination']['amount']:,.0f} financed, "
          f"${t['outstanding']['active_ob']:,.0f} outstanding, as of {as_of}")


if __name__ == "__main__":
    main()
